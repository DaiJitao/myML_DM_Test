# _*_ coding:utf-8 _*_

import numpy as np
import pandas as pd
import openpyxl as oxl

###################################### start 没用 ###############################################
class Modify_excel(object):
    """
    修改Excel的值
    """
    def __init__(self, filename):
        self.filename = filename
        self.wb = oxl.load_workbook(self.filename)
        self.ws = self.wb.active

    def modify_value(self, coord, value, savePath):
        # eg: coord:A1
        ''' 修改 '''
        self.ws.cell(coord).value = value
        self.wb.save(savePath)

    def modify_value_by_index(self, value, row_index, col_index, savePath):
        ''' 根据下标修改 '''
        self.ws.cell(row=row_index, column=col_index).value = value
        self.wb.save(savePath)

# 生成excel
def save_excel_from_matrix(file_path, save_data, file_name="excel", sheet_name="sheet_first", sheet_index=0):
    """ data 要求是numpy矩阵 """
    wb = oxl.Workbook()
    ws = wb.create_sheet(index=0, title=sheet_name)
    dim = save_data.shape
    rows = dim[0]
    cols = dim[1]
    for row_index in range(rows):
        for col_index in range(cols):
            print save_data[row_index, col_index]
            ws.cell(row=row_index + 1, column=col_index + 1).value = save_data[row_index, col_index]
            # cell的下标不能为零。从1开始
    wb.save(file_path + file_name + ".xlsx")


# 生成excel
def save_excel_from_value(file_path, value, row_index, col_index, file_name="excel", sheet_name="sheet_first",
                          sheet_index=0):
    """ 指定行列保存数据 改功能会删除原来的Excel """
    wb = oxl.Workbook()
    ws = wb.create_sheet(index=0, title=sheet_name)
    ws.cell(row=row_index, column=col_index).value = value
    wb.save(file_path + file_name + ".xlsx")
###################################### end 没用 ###############################################

def getRandomList(p_size, start=0, end=1, loop=100):
    """
     p_size:随机数的长度,个数
     loop: 100份随机数字，每份的大小为p_size
     """
    returnList = [[]] * loop
    for i in range(loop):
        prng = np.random.RandomState(12345678 + i)  # 定义局部种子
        temp = dataRandom = prng.randint(start, end, size=p_size)  # 产生start到end内的随机数
        returnList[i] = temp
    return returnList

def get_index(data):
    """找出大于零的下标, 返回list"""
    returnDict = dict()
    returnList = list()
    dim = data.shape
    rows = dim[0]
    cols = dim[1]
    count = 0
    # 下标从1开始
    for row in range(1, rows + 1):
        for col in range(row + 1, cols + 1):
            value = data[row][col]
            count += 1
            if value > 0:
                index = str(row) + "_" + str(col)
                returnDict[index] = value
                returnList.append(index)
                # print "row %s, col %s, value %s" % (row, col, value)
    print "count %s " %count
    print "rows %s" %row
    print "cols %s" %cols
    return returnList


# data = np.mat([[11,12, 13],[21,22,23]])
def get_100_excels(source_file = r"D:\temp\20180102\2010.xlsx", save_path = r"D:\temp\20180102\result_0_data"):
    data = pd.read_excel(source_file) # 源文件
    index_list = get_index(data) # 获取不为零的单元格下标
    # 获取不为零的单元格数量
    cells_len = len(index_list)
    # 提取10%
    import math
    cells_use_10 = math.ceil(cells_len * 0.1) #取整
    # 1 获取随机数字 长度为cells_len_10
    random_list = getRandomList(p_size=cells_use_10, start=0, end=cells_len)
    """ print "随机数字：", random_list
        print "random_list[99] : %s, len: %s" %(random_list[99],len(random_list[99]))
    """
    # 2 根据随机数字，使单元格变为0
    # 3 循环100次， 产生100份文件
    for i in range(0, 100):
        wb = oxl.load_workbook(source_file)  # 源文件
        ws = wb.active
        random_inner = random_list[i] # 获取使单元格变为零的下标
        for j in random_inner:
            index_row_col = index_list[j]
            row_index = int(index_row_col.split("_")[0])
            col_index = int(index_row_col.split("_")[1])
            print "index_row %s, index_col %s" %(row_index, col_index)
            ### start: 根据下标零化数据 ###
            ws.cell(row=row_index, column=col_index).value = 0
            ws.cell(row=col_index, column=row_index).value = 0
        wb.save(save_path + "\result_2010_" + str(i) + ".xlsx" )
        ###  end: 保存数据完毕###

"""读取Excel"""
# file = r"D:\temp\20180102\2010.xlsx"
# data = pd.read_excel(file)
# print "不为零的数量 %s" %len(get_index(data))
# print get_index(data)

get_100_excels()

